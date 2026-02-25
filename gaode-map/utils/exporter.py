"""
地图数据导出工具。
"""

from __future__ import annotations

import json
import logging
from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook

from modules.map_manage.schemas import MapRequest

logger = logging.getLogger(__name__)

HEADERS = [
    "名称",
    "类型编码",
    "大类",
    "中类",
    "经度",
    "纬度",
    "距离(米)",
    "线路/描述",
    "中心点",
    "半径(米)",
]

_TYPE_MAP_PATH = Path(__file__).resolve().parent.parent / "share" / "type_map.json"


def _load_type_name_map() -> dict[str, tuple[str, str]]:
    try:
        data = json.loads(_TYPE_MAP_PATH.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        logger.warning("类型配置加载失败: %s", _TYPE_MAP_PATH)
        return {}

    mapping: dict[str, tuple[str, str]] = {}
    for group in data.get("groups", []) or []:
        group_title = group.get("title", "") or ""
        for item in group.get("items", []) or []:
            key = item.get("point_type") or item.get("id") or ""
            if key:
                mapping[key] = (group_title, item.get("label", "") or "")
    return mapping


_TYPE_NAME_MAP = _load_type_name_map()


def _iter_rows(map_request: MapRequest) -> Iterable[list]:
    center_name = ""
    radius = None
    try:
        center_name = map_request.center.get("name") or ""
    except Exception:  # noqa: BLE001
        center_name = ""
    radius = map_request.radius

    for pt in map_request.points:
        lines_text = ""
        if pt.lines:
            lines_text = " / ".join(pt.lines)
        raw_type = pt.type or ""
        type_code = raw_type[5:] if raw_type.startswith("type-") else raw_type
        group_name, item_name = _TYPE_NAME_MAP.get(raw_type, ("", ""))
        yield [
            pt.name,
            type_code,
            group_name,
            item_name,
            pt.lng,
            pt.lat,
            pt.distance if pt.distance is not None else "",
            lines_text,
            center_name,
            radius,
        ]


def export_map_to_xlsx(map_request: MapRequest, map_id: int | None = None) -> tuple[str, bytes]:
    """
    将地图数据导出为 xlsx，返回 (文件名, 文件字节)。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "地图数据"
    ws.append(HEADERS)

    for row in _iter_rows(map_request):
        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"map_{map_id}_data.xlsx" if map_id is not None else "map_data.xlsx"
    logger.info("导出地图数据为 xlsx: %s", filename)
    return filename, buffer.getvalue()
